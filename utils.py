import os
import random
import csv
import io
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from models import db, StudentDetail, User, AuditLog, Notification
from config import Config

# Helper to generate student IDs
def generate_student_uid(seq_num):
    return f"MS-{datetime.now().year}-{seq_num:04d}"

# Generate Synthetic Data
def generate_synthetic_csv(file_path):
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    first_names = ["Liam", "Olivia", "Noah", "Emma", "Oliver", "Ava", "Elijah", "Charlotte", "William", "Sophia", 
                   "James", "Amelia", "Benjamin", "Isabella", "Lucas", "Mia", "Henry", "Evelyn", "Alexander", "Harper",
                   "Daniel", "Camila", "Michael", "Gianna", "Ethan", "Abigail", "Arthur", "Emily", "Sebastian", "Luna"]
    last_names = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis", "Rodriguez", "Martinez", 
                  "Hernandez", "Lopez", "Gonzalez", "Wilson", "Anderson", "Thomas", "Taylor", "Moore", "Jackson", "Martin",
                  "Lee", "Perez", "Thompson", "White", "Harris", "Sanchez", "Clark", "Ramirez", "Lewis", "Robinson"]

    depts = ["Computer Science", "Mechanical Engineering", "Electrical Engineering", "Business Analytics", "Psychology"]
    genders = ["Male", "Female", "Non-Binary"]

    rows = []
    random.seed(42)  # For reproducibility

    for i in range(1, 551):
        name = f"{random.choice(first_names)} {random.choice(last_names)}"
        gender = random.choices(genders, weights=[47, 49, 4], k=1)[0]
        age = random.randint(18, 25)
        dept = random.choice(depts)
        acad_year = random.choice([1, 2, 3, 4])
        
        # Inter-correlated features
        # We model stress based on study, sleep, screen time and attendance
        study_hours = round(random.uniform(1.5, 9.5), 1)
        sleep_hours = round(random.uniform(4.5, 9.5), 1)
        screen_time = round(random.uniform(2.0, 11.5), 1)
        attendance = round(random.uniform(60.0, 100.0), 1)
        physical_activity = round(random.uniform(0.0, 10.0), 1)
        
        # High study + low sleep + high screen + low exercise -> high fatigue & stress
        base_stress = 3.0 + (study_hours * 0.4) - (sleep_hours * 0.5) + (screen_time * 0.3) - (physical_activity * 0.15) + (100 - attendance) * 0.05
        # Normalize stress to 1-10 with random noise
        stress_level = min(10.0, max(1.0, round(base_stress + random.uniform(-1.0, 1.0), 1)))
        
        # Fatigue correlates strongly with stress and sleep
        base_fatigue = (stress_level * 0.07) + (10.0 - sleep_hours) * 0.04 + (screen_time * 0.02)
        mental_fatigue = min(1.0, max(0.0, round(base_fatigue + random.uniform(-0.05, 0.05), 2)))
        
        # Burnout Score combines indicators
        burnout_val = (stress_level * 0.35 / 10) + (mental_fatigue * 0.35) + ((10.0 - sleep_hours) * 0.15 / 10) + ((100 - attendance) * 0.15 / 100)
        burnout_score = min(1.0, max(0.0, round(burnout_val + random.uniform(-0.08, 0.08), 2)))
        
        # Categorize
        if burnout_score < 0.40:
            burnout_category = "Low"
        elif burnout_score < 0.70:
            burnout_category = "Medium"
        else:
            burnout_category = "High"

        # Introduce a few random null values & duplicates for cleaning pipeline demonstration
        # We only do this for 2% of the dataset
        if random.random() < 0.02:
            study_hours = None
        if random.random() < 0.02:
            sleep_hours = None

        student_uid = generate_student_uid(i)
        rows.append({
            'student_uid': student_uid,
            'name': name,
            'gender': gender,
            'age': age,
            'department': dept,
            'academic_year': acad_year,
            'study_hours': study_hours,
            'sleep_hours': sleep_hours,
            'stress_level': stress_level,
            'screen_time': screen_time,
            'attendance': attendance,
            'mental_fatigue': mental_fatigue,
            'physical_activity': physical_activity,
            'burnout_score': burnout_score,
            'burnout_category': burnout_category
        })

    # Let's add a few duplicates
    for _ in range(5):
        rows.append(random.choice(rows).copy())

    with open(file_path, mode='w', newline='', encoding='utf-8') as f:
        fieldnames = rows[0].keys()
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

# Data Cleaning Pipeline
def clean_and_load_dataset(csv_path):
    """
    Load data from CSV, perform cleaning (imputing missing values, removing duplicates,
    validating types), and write cleaned data into SQLite database.
    """
    if not os.path.exists(csv_path):
        generate_synthetic_csv(csv_path)

    df = pd.read_csv(csv_path)
    
    # 1. Remove duplicates
    df.drop_duplicates(subset=['student_uid'], keep='first', inplace=True)
    df.drop_duplicates(inplace=True)
    
    # 2. Handle missing values
    # Median imputations for continuous columns
    df['study_hours'] = df['study_hours'].fillna(df['study_hours'].median())
    df['sleep_hours'] = df['sleep_hours'].fillna(df['sleep_hours'].median())
    df['stress_level'] = df['stress_level'].fillna(df['stress_level'].median())
    df['screen_time'] = df['screen_time'].fillna(df['screen_time'].median())
    df['attendance'] = df['attendance'].fillna(df['attendance'].median())
    df['mental_fatigue'] = df['mental_fatigue'].fillna(df['mental_fatigue'].median())
    df['physical_activity'] = df['physical_activity'].fillna(df['physical_activity'].median())
    
    # Fill categorical / structural
    df['name'] = df['name'].fillna("Unknown Student")
    df['gender'] = df['gender'].fillna("Non-Binary")
    df['department'] = df['department'].fillna("General Studies")
    df['academic_year'] = df['academic_year'].fillna(1).astype(int)
    df['age'] = df['age'].fillna(20).astype(int)
    
    # Recalculate categories for any imputed metrics
    def recat(row):
        score = row['burnout_score']
        if pd.isna(score):
            # Recalculate
            score = (row['stress_level'] * 0.35 / 10) + (row['mental_fatigue'] * 0.35) + ((10.0 - row['sleep_hours']) * 0.15 / 10) + ((100 - row['attendance']) * 0.15 / 100)
            score = min(1.0, max(0.0, round(score, 2)))
        cat = "Low"
        if score >= 0.7:
            cat = "High"
        elif score >= 0.4:
            cat = "Medium"
        return score, cat
        
    res = df.apply(recat, axis=1)
    df['burnout_score'] = [r[0] for r in res]
    df['burnout_category'] = [r[1] for r in res]
    
    # Seed into database
    # Clear existing details
    db.session.query(StudentDetail).delete()
    
    # Bulk insert
    student_details = []
    for idx, row in df.iterrows():
        student = StudentDetail(
            student_uid=row['student_uid'],
            name=row['name'],
            gender=row['gender'],
            age=int(row['age']),
            department=row['department'],
            academic_year=int(row['academic_year']),
            study_hours=float(row['study_hours']),
            sleep_hours=float(row['sleep_hours']),
            stress_level=float(row['stress_level']),
            screen_time=float(row['screen_time']),
            attendance=float(row['attendance']),
            mental_fatigue=float(row['mental_fatigue']),
            physical_activity=float(row['physical_activity']),
            burnout_score=float(row['burnout_score']),
            burnout_category=row['burnout_category']
        )
        student_details.append(student)
        
    db.session.bulk_save_objects(student_details)
    db.session.commit()
    
    # Create a system notification about data readiness
    notif = Notification(
        title="Burnout Dataset Synced",
        message=f"Dataset successfully cleaned and loaded. Loaded {len(student_details)} active student records into database tables.",
        type="success"
    )
    db.session.add(notif)
    db.session.commit()

# Seed Auth accounts
def seed_default_users():
    # Admin
    if not User.query.filter_by(email="admin@mindspace.edu").first():
        admin = User(
            email="admin@mindspace.edu",
            name="Dr. Sarah Jenkins",
            role="admin",
            department="Academic Affairs",
            semester="N/A"
        )
        admin.set_password("AdminPass123!")
        db.session.add(admin)
        
    # Student
    if not User.query.filter_by(email="student@mindspace.edu").first():
        student = User(
            email="student@mindspace.edu",
            name="Alex Rivera",
            role="student",
            department="Computer Science",
            semester="Semester 5"
        )
        student.set_password("StudentPass123!")
        db.session.add(student)
        
    db.session.commit()

# EXPORT TO CSV
def export_students_csv(queryset):
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow([
        'Student ID', 'Name', 'Gender', 'Age', 'Department', 'Academic Year', 
        'Study Hours/Day', 'Sleep Hours/Day', 'Stress Level (1-10)', 'Screen Time/Day', 
        'Attendance (%)', 'Mental Fatigue (0-1)', 'Physical Activity (hrs/wk)', 'Burnout Score (0-1)', 'Risk Category'
    ])
    
    for s in queryset:
        writer.writerow([
            s.student_uid, s.name, s.gender, s.age, s.department, s.academic_year,
            s.study_hours, s.sleep_hours, s.stress_level, s.screen_time,
            s.attendance, s.mental_fatigue, s.physical_activity, s.burnout_score, s.burnout_category
        ])
        
    return output.getvalue()

# EXPORT TO EXCEL
def export_students_excel(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "MindSpace Student Data"
    
    # Styles
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="5B8DEF", end_color="5B8DEF", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border_side = Side(style='thin', color="DDDDDD")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    headers = [
        'Student ID', 'Name', 'Gender', 'Age', 'Department', 'Academic Year', 
        'Study Hours/Day', 'Sleep Hours/Day', 'Stress Level (1-10)', 'Screen Time/Day', 
        'Attendance (%)', 'Mental Fatigue (0-1)', 'Physical Activity (hrs/wk)', 'Burnout Score (0-1)', 'Risk Category'
    ]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    for s in queryset:
        row = [
            s.student_uid, s.name, s.gender, s.age, s.department, s.academic_year,
            s.study_hours, s.sleep_hours, s.stress_level, s.screen_time,
            s.attendance, s.mental_fatigue, s.physical_activity, s.burnout_score, s.burnout_category
        ]
        ws.append(row)
        
    # Apply styling to cells
    for r_idx in range(2, ws.max_row + 1):
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = border
            if c_idx in [1, 3, 4, 6]:
                cell.alignment = center_align

    # Auto-adjust column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 10)
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

# EXPORT TO PDF
def export_students_pdf(queryset, title_name="Student Burnout Report"):
    """
    Generate a clean PDF report with ReportLab.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=40, bottomMargin=40)
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor('#5B8DEF'),
        spaceAfter=15
    )
    
    meta_style = ParagraphStyle(
        'DocMeta',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#666666'),
        spaceAfter=25
    )
    
    body_style = ParagraphStyle(
        'DocBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=12
    )

    header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white
    )
    
    elements = []
    
    # Header Section
    elements.append(Paragraph(f"MindSpace System Analytics", title_style))
    elements.append(Paragraph(f"Report: {title_name} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", meta_style))
    elements.append(Spacer(1, 10))
    
    # Create Report Table
    table_data = [[
        Paragraph('ID', header_style),
        Paragraph('Name', header_style),
        Paragraph('Dept', header_style),
        Paragraph('Sleep', header_style),
        Paragraph('Study', header_style),
        Paragraph('Stress', header_style),
        Paragraph('Attend', header_style),
        Paragraph('Fatigue', header_style),
        Paragraph('Risk', header_style)
    ]]
    
    # Let's cap the PDF rows at 50 in the print table to avoid a massive file size
    for s in list(queryset)[:80]:
        table_data.append([
            Paragraph(s.student_uid, body_style),
            Paragraph(s.name[:18], body_style),
            Paragraph(s.department[:15], body_style),
            Paragraph(f"{s.sleep_hours:.1f}", body_style),
            Paragraph(f"{s.study_hours:.1f}", body_style),
            Paragraph(f"{s.stress_level:.1f}", body_style),
            Paragraph(f"{s.attendance:.1f}%", body_style),
            Paragraph(f"{s.mental_fatigue:.2f}", body_style),
            Paragraph(f"<b>{s.burnout_category}</b>", ParagraphStyle('RiskCell', parent=body_style, textColor=colors.HexColor('#EF5350') if s.burnout_category == 'High' else colors.HexColor('#F4B400') if s.burnout_category == 'Medium' else colors.HexColor('#4CAF50')))
        ])
        
    table = Table(table_data, colWidths=[65, 85, 85, 40, 40, 40, 45, 45, 55])
    
    # Table Styling
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5B8DEF')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
    ])
    
    # Add alternating row background colors
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F9FAFC'))
            
    table.setStyle(t_style)
    elements.append(table)
    
    # Footer Notice
    elements.append(Spacer(1, 20))
    if len(list(queryset)) > 80:
        elements.append(Paragraph(f"<i>* Showing top 80 matching rows. Out of {len(list(queryset))} total matching records.</i>", meta_style))
        
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
